#!/usr/bin/env python3
"""
分析测试结果 — 单shape / 多shape 通用
产出到 output/：
  - best_kernels_report.xlsx   每个shape最优kernel汇总
  - *.py                        所有最优kernel代码
"""
import paramiko
import json
import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), 'output')

with open(os.path.join(SCRIPT_DIR, 'server_config.json')) as f:
    config = json.load(f)

CONTAINER = config['docker_container']
RESULTS = '/root/MyAICode/batch_verification/results'
CODES = '/root/MyAICode/batch_verification/codes'

def ssh_cmd(ssh, cmd):
    stdin, stdout, stderr = ssh.exec_command(f'docker exec {CONTAINER} bash -c "{cmd}"', timeout=30)
    return stdout.read().decode('utf-8'), stderr.read().decode('utf-8')

def get_summary(ssh):
    out, _ = ssh_cmd(ssh, f'cat {RESULTS}/summary.json 2>/dev/null')
    if not out.strip():
        return None
    return json.loads(out)

def get_shape_summary(ssh, shape):
    filename = f'summary_M{shape["M"]}_N{shape["N"]}_K{shape["K"]}.json'
    out, _ = ssh_cmd(ssh, f'cat {RESULTS}/{filename} 2>/dev/null')
    if not out.strip():
        return None
    return json.loads(out)

def download_kernel(ssh, filename):
    cmd = f'find {CODES} -name "{filename}" -exec cat {{}} \\;'
    out, _ = ssh_cmd(ssh, cmd)
    if not out.strip():
        return None
    local = os.path.join(OUTPUT_DIR, filename)
    with open(local, 'w', encoding='utf-8') as f:
        f.write(out)
    return local

def generate_excel(best_kernels):
    path = os.path.join(OUTPUT_DIR, 'best_kernels_report.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "Best Kernels"

    headers = ["Shape", "M", "N", "K", "Task Duration (us)", "Passed", "Kernel File", "Kernel Name"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in best_kernels:
        s = item['shape']
        k = item['kernel']
        ws.append([
            f"M={s['M']}, N={s['N']}, K={s['K']}" if s else "default",
            s['M'] if s else "", s['N'] if s else "", s['K'] if s else "",
            k['duration'], k['passed'], k['file'], k['kernel']
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)
    return path

ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(hostname=config['ip'], username='root', key_filename=config['ssh_key_path'], timeout=30)

try:
    summary = get_summary(ssh)
    if not summary:
        print('No summary.json found on server')
        exit(1)

    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    total_shapes = summary.get('total_shapes', 1)
    all_results = summary.get('results', [])

    best_kernels = []

    if total_shapes == 1 and all_results:
        shape = all_results[0].get('shape') if all_results else None
        passed = [r for r in all_results if r.get('passed')]
        passed.sort(key=lambda r: r['duration'])
        best = passed[0] if passed else None
        if best:
            best_kernels.append({'shape': shape, 'kernel': best})

        print("=" * 95)
        print("测试结果统计")
        print("=" * 95)
        total = summary.get('total_tests', len(all_results))
        p = summary.get('total_passed', len(passed))
        f = summary.get('total_failed', total - p)
        print(f"总共测试kernel数量: {total}")
        print(f"通过测试kernel数量: {p}")
        print(f"失败测试kernel数量: {f}")
        print(f"通过率: {p/total*100:.2f}%" if total else "")
        print()

        print("=" * 95)
        print("最优Kernel")
        print("=" * 95)
        print(f"{'Rank':<6} {'File':<70} {'Duration(us)':<15}")
        print("=" * 95)
        for i, r in enumerate(passed[:1], 1):
            print(f"{i:<6} {r['file']:<70} {r['duration']:<15.2f}")
        print("=" * 95)
    else:
        for shape in summary.get('shapes', []):
            if shape is None:
                continue
            ss = get_shape_summary(ssh, shape)
            if not ss:
                continue
            results = ss.get('results', [])
            passed = [r for r in results if r.get('passed')]
            if not passed:
                continue
            passed.sort(key=lambda r: r['duration'])
            best_kernels.append({'shape': shape, 'kernel': passed[0]})

        print("=" * 95)
        print("多Shape测试结果分析")
        print("=" * 95)
        print(f"总Shape数: {len(best_kernels)}")
        print(f"总测试数: {summary.get('total_tests','?')}")
        print(f"总通过数: {summary.get('total_passed','?')}")
        print(f"总失败数: {summary.get('total_failed','?')}")

    if not best_kernels:
        print("\n没有找到通过测试的kernel")
        exit(0)

    print()
    print("=" * 95)
    print("各Shape最优Kernel")
    print("=" * 95)
    for item in best_kernels:
        s = item['shape']
        k = item['kernel']
        print(f"  M={s['M']}, N={s['N']}, K={s['K']}  |  {k['file']}  |  {k['duration']:.2f} us")

    print(f"\n正在从服务器下载最优kernel代码到 {OUTPUT_DIR} ...")
    for item in best_kernels:
        local = download_kernel(ssh, item['kernel']['file'])
        if local:
            print(f"  ✓ {item['kernel']['file']}")
        else:
            print(f"  ✗ {item['kernel']['file']} (failed)")

    excel = generate_excel(best_kernels)
    print(f"\n✓ Excel报告: {excel}")
    print("✓ 完成")

finally:
    ssh.close()
