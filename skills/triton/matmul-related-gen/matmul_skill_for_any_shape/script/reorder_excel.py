#!/usr/bin/env python3
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read merged_results_all.xlsx to get shape order
print("Reading merged_results_all.xlsx...")
wb1 = load_workbook(r'D:\AI-Triton\AICoding\matmul_skill_for_any_shape\output\merged_results_all.xlsx')
ws1 = wb1.active

merged_shapes = []
for row_idx, row in enumerate(ws1.iter_rows(min_row=2, values_only=True), 2):
    if len(row) >= 3:
        m, n, k = row[0], row[1], row[2]
        merged_shapes.append((m, n, k))
        print("  Shape %d: M=%s, N=%s, K=%s" % (len(merged_shapes), m, n, k))

# Read best_kernels_report.xlsx
print("\nReading best_kernels_report.xlsx...")
wb2 = load_workbook(r'D:\AI-Triton\AICoding\matmul_skill_for_any_shape\output\best_kernels_report.xlsx')
ws2 = wb2.active

best_kernels_data = {}
for row_idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), 2):
    if len(row) >= 7:
        m, n, k = row[1], row[2], row[3]
        key = (m, n, k)
        best_kernels_data[key] = row
        print("  Found: M=%s, N=%s, K=%s" % (m, n, k))

# Create new Excel file
print("\nCreating reordered Excel...")
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "Best Kernels"

# Add header
headers = ["Shape", "M", "N", "K", "Task Duration (us)", "Passed", "Kernel File", "Kernel Name"]
ws_new.append(headers)

# Apply header styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col_idx, cell in enumerate(ws_new[1], 1):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data in merged_results_all order
count = 0
for (m, n, k) in merged_shapes:
    key = (m, n, k)
    if key in best_kernels_data:
        row = best_kernels_data[key]
        ws_new.append(row)
        count += 1
        print("  Added: M=%s, N=%s, K=%s" % (m, n, k))
    else:
        print("  Warning: M=%s, N=%s, K=%s not found in best_kernels_report" % (m, n, k))

# Adjust column widths
for col in ws_new.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_new.column_dimensions[column].width = adjusted_width

# Save
output_path = r'D:\AI-Triton\AICoding\matmul_skill_for_any_shape\output\best_kernels_report_reordered.xlsx'
wb_new.save(output_path)
print("\nSuccess! Reordered Excel saved to: %s" % output_path)
print("Total shapes added: %d/%d" % (count, len(merged_shapes)))
